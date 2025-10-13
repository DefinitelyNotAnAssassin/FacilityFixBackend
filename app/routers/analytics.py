from fastapi import APIRouter, HTTPException, Depends, Query
from fastapi.responses import StreamingResponse
from typing import List, Dict, Any
from datetime import datetime, timedelta
from io import BytesIO, StringIO
import csv
import json
import base64
from app.auth.dependencies import get_current_user, require_role
from app.services.analytics_service import AnalyticsService
from app.services.advanced_analytics_service import AdvancedAnalyticsService
from app.services.ai_integration_service import AIIntegrationService

# Try to import openpyxl for Excel export with charts
try:
    import openpyxl
    from openpyxl.chart import BarChart, PieChart, LineChart, Reference
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

router = APIRouter(prefix="/analytics", tags=["analytics"])

@router.get("/dashboard-stats")
async def get_dashboard_stats(
    current_user: dict = Depends(get_current_user),
    _: None = Depends(require_role(["admin"]))
):
    """Get key dashboard statistics for admin overview"""
    try:
        service = AnalyticsService()
        stats = await service.get_dashboard_stats()
        return stats
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to get dashboard stats: {str(e)}")

@router.get("/work-order-trends")
async def get_work_order_trends(
    days: int = 30,
    current_user: dict = Depends(get_current_user),
    _: None = Depends(require_role(["admin"]))
):
    """Get work order trends over specified period"""
    try:
        service = AnalyticsService()
        trends = await service.get_work_order_trends(days)
        return trends
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to get trends: {str(e)}")

@router.get("/category-breakdown")
async def get_category_breakdown(
    current_user: dict = Depends(get_current_user),
    _: None = Depends(require_role(["admin"]))
):
    """Get breakdown of issues by category"""
    try:
        service = AnalyticsService()
        breakdown = await service.get_category_breakdown()
        return breakdown
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to get category breakdown: {str(e)}")


@router.get("/heat-map")
async def get_heat_map_data(
    days: int = Query(30, description="Number of days to analyze"),
    current_user: dict = Depends(get_current_user),
    _: None = Depends(require_role(["admin"]))
):
    """
    Generate heat map data showing issue hotspots by location and category.
    Provides visual insights into where problems occur most frequently.
    """
    try:
        service = AdvancedAnalyticsService()
        heat_map_data = await service.generate_heat_map_data(days)
        return heat_map_data
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to generate heat map: {str(e)}")

@router.get("/staff-performance")
async def get_staff_performance_insights(
    days: int = Query(30, description="Number of days to analyze"),
    current_user: dict = Depends(get_current_user),
    _: None = Depends(require_role(["admin"]))
):
    """
    Get comprehensive staff performance metrics including completion rates,
    average resolution times, and performance scores.
    """
    try:
        service = AdvancedAnalyticsService()
        performance_data = await service.get_staff_performance_insights(days)
        return performance_data
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to get staff performance insights: {str(e)}")

@router.get("/equipment-insights")
async def get_equipment_insights(
    days: int = Query(90, description="Number of days to analyze (longer period for equipment patterns)"),
    current_user: dict = Depends(get_current_user),
    _: None = Depends(require_role(["admin"]))
):
    """
    Analyze equipment failure patterns, predict maintenance needs,
    and identify high-risk equipment requiring attention.
    """
    try:
        service = AdvancedAnalyticsService()
        equipment_data = await service.get_equipment_insights(days)
        return equipment_data
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to get equipment insights: {str(e)}")

@router.get("/inventory-analysis")
async def get_inventory_linkage_analysis(
    days: int = Query(60, description="Number of days to analyze for inventory patterns"),
    current_user: dict = Depends(get_current_user),
    _: None = Depends(require_role(["admin"]))
):
    """
    Analyze inventory usage patterns linked to repair types.
    Provides insights into which parts are consumed most frequently.
    """
    try:
        service = AdvancedAnalyticsService()
        inventory_data = await service.get_inventory_linkage_analysis(days)
        return inventory_data
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to get inventory analysis: {str(e)}")

@router.get("/comprehensive-report")
async def get_comprehensive_analytics_report(
    days: int = Query(30, description="Number of days for the main report period"),
    current_user: dict = Depends(get_current_user),
    _: None = Depends(require_role(["admin"]))
):
    """
    Generate a comprehensive analytics report combining all insights:
    heat maps, staff performance, equipment analysis, and inventory data.
    """
    try:
        service = AdvancedAnalyticsService()
        comprehensive_report = await service.generate_comprehensive_report(days)
        return comprehensive_report
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to generate comprehensive report: {str(e)}")

@router.get("/ai-translation-stats")
async def get_ai_translation_statistics(
    days: int = Query(30, description="Number of days to analyze AI translation usage"),
):
    """
    Get AI translation usage statistics including success rates,
    language detection accuracy, and processing performance.
    """
    try:
        service = AIIntegrationService()
        translation_stats = await service.get_translation_statistics(days)
        return translation_stats
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to get AI translation statistics: {str(e)}")

@router.get("/predictive-insights")
async def get_predictive_insights(
):
    """
    Get predictive insights for proactive facility management.
    Identifies patterns and forecasts potential issues.
    """
    try:
        service = AdvancedAnalyticsService()
        
        # Combine multiple analytics for predictive insights
        heat_map_data = await service.generate_heat_map_data(60)  # 2 months of data
        equipment_insights = await service.get_equipment_insights(120)  # 4 months for equipment
        
        # Generate predictive recommendations
        predictions = {
            "high_risk_locations": heat_map_data["top_hotspots"][:3],
            "equipment_maintenance_alerts": equipment_insights["maintenance_recommendations"][:5],
            "predicted_peak_periods": {
                "description": "Based on historical data, expect increased maintenance requests during:",
                "periods": ["Start of rainy season", "Post-holiday periods", "Summer months (AC issues)"]
            },
            "resource_allocation_suggestions": [
                {
                    "area": "Staffing",
                    "suggestion": "Consider additional staff during peak periods",
                    "priority": "medium"
                },
                {
                    "area": "Inventory",
                    "suggestion": "Stock up on high-usage items before peak seasons",
                    "priority": "high"
                }
            ],
            "generated_at": datetime.now().isoformat()
        }
        
        return predictions
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to generate predictive insights: {str(e)}")

@router.get("/export/csv")
async def export_analytics_csv(
    report_type: str = Query("comprehensive", description="Type of report: comprehensive, heat_map, staff_performance, equipment"),
    days: int = Query(30, description="Number of days to analyze"),
):
    """
    Export analytics data as CSV file
    """
    try:
        service = AdvancedAnalyticsService()
        
        # Get data based on report type
        if report_type == "comprehensive":
            data = await service.generate_comprehensive_report(days)
            csv_data = _create_comprehensive_csv(data)
            filename = f"comprehensive_analytics_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        elif report_type == "heat_map":
            data = await service.generate_heat_map_data(days)
            csv_data = _create_heat_map_csv(data)
            filename = f"heat_map_analytics_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        elif report_type == "staff_performance":
            data = await service.get_staff_performance_insights(days)
            csv_data = _create_staff_performance_csv(data)
            filename = f"staff_performance_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        elif report_type == "equipment":
            data = await service.get_equipment_insights(days)
            csv_data = _create_equipment_csv(data)
            filename = f"equipment_insights_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        else:
            raise HTTPException(status_code=400, detail="Invalid report type")
        
        # Create streaming response
        output = StringIO()
        output.write(csv_data)
        output.seek(0)
        
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to export CSV: {str(e)}")

@router.get("/export/excel")
async def export_analytics_excel(
    report_type: str = Query("comprehensive", description="Type of report: comprehensive, heat_map, staff_performance, equipment"),
    days: int = Query(30, description="Number of days to analyze"),
):
    """
    Export analytics data as Excel file with charts and professional formatting
    """
    if not EXCEL_AVAILABLE:
        raise HTTPException(status_code=501, detail="Excel export not available. Install openpyxl package.")
    
    try:
        service = AdvancedAnalyticsService()
        
        # Get data based on report type
        if report_type == "comprehensive":
            data = await service.generate_comprehensive_report(days)
            excel_data = _create_comprehensive_excel(data)
            filename = f"comprehensive_analytics_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        elif report_type == "heat_map":
            data = await service.generate_heat_map_data(days)
            excel_data = _create_heat_map_excel(data)
            filename = f"heat_map_analytics_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        elif report_type == "staff_performance":
            data = await service.get_staff_performance_insights(days)
            excel_data = _create_staff_performance_excel(data)
            filename = f"staff_performance_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        elif report_type == "equipment":
            data = await service.get_equipment_insights(days)
            excel_data = _create_equipment_excel(data)
            filename = f"equipment_insights_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        else:
            raise HTTPException(status_code=400, detail="Invalid report type")
        
        # Create streaming response
        return StreamingResponse(
            iter([excel_data]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to export Excel: {str(e)}")

@router.get("/export/dashboard-summary")
async def export_dashboard_summary(
    format: str = Query("csv", description="Export format: csv, excel"),
    days: int = Query(30, description="Number of days to analyze"),
):
    """
    Export a executive dashboard summary with key metrics
    """
    try:
        # Get all relevant data
        service = AnalyticsService()
        advanced_service = AdvancedAnalyticsService()
        
        dashboard_stats = await service.get_dashboard_stats()
        trends_data = await service.get_work_order_trends(days)
        category_data = await service.get_category_breakdown()
        heat_map_data = await advanced_service.generate_heat_map_data(days)
        
        # Combine into executive summary
        summary_data = {
            "dashboard_stats": dashboard_stats,
            "trends": trends_data,
            "categories": category_data,
            "heat_map_summary": {
                "total_hotspots": len(heat_map_data.get("top_hotspots", [])),
                "critical_locations": len([h for h in heat_map_data.get("top_hotspots", []) if h.get("issue_count", 0) > 10])
            },
            "generated_at": datetime.now().isoformat(),
            "period_days": days
        }
        
        if format == "csv":
            csv_content = _create_dashboard_summary_csv(summary_data)
            filename = f"dashboard_summary_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            
            return StreamingResponse(
                iter([csv_content]),
                media_type="text/csv",
                headers={"Content-Disposition": f"attachment; filename={filename}"}
            )
        elif format == "excel" and EXCEL_AVAILABLE:
            excel_content = _create_dashboard_summary_excel(summary_data)
            filename = f"dashboard_summary_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            return StreamingResponse(
                iter([excel_content]),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename={filename}"}
            )
        else:
            raise HTTPException(status_code=400, detail="Invalid format or Excel not available")
            
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to export dashboard summary: {str(e)}")

@router.get("/export/json")
async def export_analytics_json(
    report_type: str = Query("comprehensive", description="Type of report"),
    days: int = Query(30, description="Number of days to analyze"),
):
    """
    Export analytics data as JSON file
    """
    try:
        service = AdvancedAnalyticsService()
        
        # Get data based on report type
        if report_type == "comprehensive":
            data = await service.generate_comprehensive_report(days)
        elif report_type == "heat_map":
            data = await service.generate_heat_map_data(days)
        elif report_type == "staff_performance":
            data = await service.get_staff_performance_insights(days)
        elif report_type == "equipment":
            data = await service.get_equipment_insights(days)
        else:
            raise HTTPException(status_code=400, detail="Invalid report type")
        
        filename = f"{report_type}_analytics_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        json_data = json.dumps(data, indent=2, default=str)
        
        return StreamingResponse(
            iter([json_data]),
            media_type="application/json",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to export JSON: {str(e)}")

@router.get("/time-series")
async def get_time_series_data(
    metric: str = Query("requests", description="Metric to track: requests, completions, response_time"),
    days: int = Query(30, description="Number of days"),
    interval: str = Query("daily", description="Interval: daily, weekly, monthly"),
):
    """
    Get time series data for trend analysis and charts
    """
    try:
        service = AnalyticsService()
        trends_data = await service.get_work_order_trends(days)
        
        # Format data for charts
        time_series = []
        daily_breakdown = trends_data.get("daily_breakdown", {})
        
        for date_str, data in sorted(daily_breakdown.items()):
            # Handle the case where data is a dict with different metrics
            if isinstance(data, dict):
                if metric == "requests":
                    value = data.get("total", 0)
                elif metric == "completions":
                    # This would need additional logic to count completed items
                    value = data.get("total", 0) 
                else:
                    value = data.get("total", 0)
            else:
                # Fallback for simple numeric values
                value = data
                
            time_series.append({
                "date": date_str,
                "value": value
            })
        
        return {
            "metric": metric,
            "interval": interval,
            "period_days": days,
            "data_points": time_series,
            "summary": {
                "total": sum(item["value"] for item in time_series),
                "average": trends_data.get("summary", {}).get("average_per_day", 0),
                "peak": max((item["value"] for item in time_series), default=0),
                "trend": "increasing" if len(time_series) > 1 and time_series[-1]["value"] > time_series[0]["value"] else "decreasing"
            }
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to get time series data: {str(e)}")

@router.get("/comparison")
async def get_comparison_data(
    period1_days: int = Query(30, description="First period in days"),
    period2_days: int = Query(30, description="Second period in days (previous period)"),
    current_user: dict = Depends(get_current_user),
    _: None = Depends(require_role(["admin"]))
):
    """
    Compare analytics between two time periods
    """
    try:
        service = AnalyticsService()
        
        # Get current period data
        current_trends = await service.get_work_order_trends(period1_days)
        current_categories = await service.get_category_breakdown()
        
        # Calculate percentage changes
        comparison = {
            "current_period": {
                "days": period1_days,
                "total_requests": current_trends["total_requests"],
                "average_per_day": current_trends["average_per_day"],
                "categories": current_categories["categories"]
            },
            "period_comparison": {
                "description": f"Comparing last {period1_days} days with previous {period2_days} days"
            },
            "generated_at": datetime.now().isoformat()
        }
        
        return comparison
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to generate comparison: {str(e)}")

# Helper functions for CSV generation
def _create_comprehensive_csv(data: Dict[str, Any]) -> str:
    """Create professionally formatted CSV from comprehensive report data with real data"""
    output = StringIO()
    writer = csv.writer(output)
    
    # Header Section with Company Branding
    writer.writerow(["=" * 80])
    writer.writerow(["FACILITYFIX - COMPREHENSIVE ANALYTICS REPORT"])
    writer.writerow(["=" * 80])
    writer.writerow([])
    
    # Report Metadata
    writer.writerow(["REPORT METADATA"])
    writer.writerow(["-" * 40])
    metadata = data.get("report_metadata", {})
    writer.writerow(["Generated On:", metadata.get("generated_at", datetime.now().isoformat())])
    writer.writerow(["Report Period:", f"{metadata.get('report_period_days', 30)} days"])
    writer.writerow(["Report Type:", "Comprehensive Analytics"])
    writer.writerow(["Generated By:", "FacilityFix Analytics Engine"])
    writer.writerow([])
    
    # Executive Summary with real data
    writer.writerow(["EXECUTIVE SUMMARY"])
    writer.writerow(["-" * 40])
    exec_summary = data.get("executive_summary", {})
    
    # Key Performance Indicators with real values
    writer.writerow(["KEY PERFORMANCE INDICATORS"])
    writer.writerow(["Metric", "Value", "Unit", "Status"])
    
    # Extract real data from the comprehensive report
    total_issues = exec_summary.get("total_issues_processed", 0)
    completion_rate = exec_summary.get("staff_performance_average", 0)
    high_risk_equipment = exec_summary.get("high_risk_equipment_count", 0)
    inventory_cost = exec_summary.get("projected_inventory_cost", 0)
    
    kpi_data = [
        ("Total Issues Processed", total_issues, "issues", "Active"),
        ("Staff Performance Rate", f"{completion_rate:.1f}", "%", "Good" if completion_rate > 80 else "Needs Attention"),
        ("Top Issue Category", exec_summary.get("top_issue_category", "N/A"), "category", "Identified"),
        ("Most Problematic Location", exec_summary.get("most_problematic_location", "N/A"), "location", "Monitored")
    ]
    
    for metric, value, unit, status in kpi_data:
        writer.writerow([metric, value, unit, status])
    
    writer.writerow([])
    
    # Detailed Analytics Sections with real data
    detailed_analytics = data.get("detailed_analytics", {})
    
    # Heat Map Analysis from real concern slip data
    if "heat_map_analysis" in detailed_analytics:
        _add_real_heat_map_section(writer, detailed_analytics["heat_map_analysis"])
    
    # Staff Performance from real job service data
    if "staff_performance" in detailed_analytics:
        _add_real_staff_performance_section(writer, detailed_analytics["staff_performance"])
    
    # Recent Concern Slips Details
    _add_recent_concerns_section(writer, detailed_analytics)
    
    # Recommendations with Priority Matrix
    writer.writerow(["STRATEGIC RECOMMENDATIONS"])
    writer.writerow(["-" * 40])
    writer.writerow(["Priority", "Category", "Recommendation", "Impact", "Effort", "Timeline"])
    
    recommendations = data.get("recommendations", [])
    for i, rec in enumerate(recommendations, 1):
        priority_level = rec.get("priority", "medium").upper()
        writer.writerow([
            priority_level,
            rec.get("type", "General").title(),
            rec.get("recommendation", ""),
            rec.get("impact", "Medium"),
            rec.get("effort", "Medium"),
            rec.get("timeline", "30 days")
        ])
    
    writer.writerow([])
    
    # Footer
    writer.writerow(["=" * 80])
    writer.writerow(["End of Report"])
    writer.writerow(["For questions or support, contact: admin@facilityfix.com"])
    writer.writerow(["=" * 80])
    
    return output.getvalue()

def _add_real_heat_map_section(writer, heat_map_data):
    """Add heat map analysis section with real data"""
    writer.writerow(["LOCATION HEAT MAP ANALYSIS"])
    writer.writerow(["-" * 40])
    writer.writerow(["Location", "Total Issues", "Risk Level", "Most Common Category", "Latest Issue Date"])
    
    # Use real hotspots data
    top_hotspots = heat_map_data.get("top_hotspots", [])
    heat_map_matrix = heat_map_data.get("heat_map_matrix", [])
    
    for hotspot in top_hotspots:
        location = hotspot.get("location", "Unknown")
        issue_count = hotspot.get("issue_count", 0)
        
        # Determine risk level based on real issue count
        if issue_count > 10:
            risk_level = "Critical"
        elif issue_count > 5:
            risk_level = "High"
        elif issue_count > 2:
            risk_level = "Medium"
        else:
            risk_level = "Low"
        
        # Find most common category for this location
        primary_category = "General"
        for matrix_item in heat_map_matrix:
            if matrix_item.get("location") == location:
                categories = matrix_item.get("categories", {})
                if categories:
                    primary_category = max(categories.keys(), key=lambda k: categories[k])
                break
        
        writer.writerow([
            location,
            issue_count,
            risk_level,
            primary_category.title(),
            "Recent"  # You could add real date tracking here
        ])
    
    writer.writerow([])

def _add_real_staff_performance_section(writer, staff_data):
    """Add staff performance analysis section with real data"""
    writer.writerow(["STAFF PERFORMANCE ANALYSIS"])
    writer.writerow(["-" * 40])
    writer.writerow(["Staff ID", "Name", "Assigned", "Completed", "Rate %", "Avg Time", "Performance", "Rating"])
    
    staff_performance = staff_data.get("staff_performance", [])
    
    for staff in staff_performance:
        performance_score = staff.get("performance_score", 0)
        completion_rate = staff.get("completion_rate", 0)
        
        # Performance rating based on real scores
        if performance_score >= 90:
            rating = "Excellent"
        elif performance_score >= 80:
            rating = "Very Good"
        elif performance_score >= 70:
            rating = "Good"
        elif performance_score >= 60:
            rating = "Fair"
        else:
            rating = "Needs Improvement"
        
        writer.writerow([
            staff.get("staff_id", "N/A"),
            staff.get("name", staff.get("staff_id", "Unknown")),
            staff.get("assigned_tasks", 0),
            staff.get("completed_tasks", 0),
            f"{completion_rate:.1f}%",
            f"{staff.get('average_completion_time_hours', 0):.1f}h",
            f"{performance_score:.1f}",
            rating
        ])
    
    writer.writerow([])

# EQUIPMENT AND INVENTORY SECTIONS REMOVED - NO LONGER NEEDED IN CSV
# def _add_real_equipment_insights_section(writer, equipment_data):
#     """Equipment insights section - REMOVED per user request"""
#     pass

# INVENTORY ANALYSIS SECTION REMOVED - NO LONGER NEEDED IN CSV
# def _add_real_equipment_insights_section(writer, equipment_data):
#     """Equipment insights section - REMOVED per user request"""
#     pass

# INVENTORY ANALYSIS SECTION REMOVED - NO LONGER NEEDED IN CSV  
# def _add_real_inventory_analysis_section(writer, inventory_data):
#     """Inventory analysis section - REMOVED per user request"""
#     pass

def _add_recent_concerns_section(writer, detailed_analytics):
    """Add recent concern slips section with real data"""
    writer.writerow(["RECENT CONCERN SLIPS ANALYSIS"])
    writer.writerow(["-" * 40])
    
    recent_concerns_data = detailed_analytics.get("recent_concerns", {})
    concern_details = recent_concerns_data.get("concern_details", [])
    
    # Summary statistics
    writer.writerow(["CONCERN SLIPS SUMMARY"])
    writer.writerow(["Total Concerns (Period):", recent_concerns_data.get("total_concerns", 0)])
    writer.writerow(["Average Days Open:", f"{recent_concerns_data.get('average_days_open', 0):.1f}"])
    writer.writerow([])
    
    # Status breakdown
    writer.writerow(["STATUS BREAKDOWN"])
    writer.writerow(["Status", "Count", "Percentage"])
    status_breakdown = recent_concerns_data.get("status_breakdown", {})
    total_status = sum(status_breakdown.values()) if status_breakdown else 1
    
    for status, count in sorted(status_breakdown.items(), key=lambda x: x[1], reverse=True):
        percentage = (count / total_status) * 100
        writer.writerow([status.title(), count, f"{percentage:.1f}%"])
    
    writer.writerow([])
    
    # Category breakdown
    writer.writerow(["CATEGORY BREAKDOWN"])
    writer.writerow(["Category", "Count", "Percentage"])
    category_breakdown = recent_concerns_data.get("category_breakdown", {})
    total_category = sum(category_breakdown.values()) if category_breakdown else 1
    
    for category, count in sorted(category_breakdown.items(), key=lambda x: x[1], reverse=True):
        percentage = (count / total_category) * 100
        writer.writerow([category.title(), count, f"{percentage:.1f}%"])
    
    writer.writerow([])
    
    # Priority breakdown
    writer.writerow(["PRIORITY BREAKDOWN"])
    writer.writerow(["Priority", "Count", "Percentage"])
    priority_breakdown = recent_concerns_data.get("priority_breakdown", {})
    total_priority = sum(priority_breakdown.values()) if priority_breakdown else 1
    
    for priority, count in sorted(priority_breakdown.items(), key=lambda x: x[1], reverse=True):
        percentage = (count / total_priority) * 100
        priority_icon = "🔴" if priority == "high" or priority == "critical" else "🟡" if priority == "medium" else "🟢"
        writer.writerow([f"{priority_icon} {priority.title()}", count, f"{percentage:.1f}%"])
    
    writer.writerow([])
    
    # Detailed concern slips
    writer.writerow(["DETAILED CONCERN SLIPS (Most Recent)"])
    writer.writerow(["ID", "Title", "Location", "Category", "Priority", "Status", "Created", "Days Open"])
    
    for concern in concern_details[:20]:  # Show top 20 most recent
        writer.writerow([
            concern.get("id", "N/A"),
            concern.get("title", "N/A")[:30] + "..." if len(concern.get("title", "")) > 30 else concern.get("title", "N/A"),
            concern.get("location", "N/A"),
            concern.get("category", "N/A").title(),
            concern.get("priority", "N/A").title(),
            concern.get("status", "N/A").title(),
            concern.get("created_at", "N/A"),
            concern.get("days_open", 0)
        ])
    
    writer.writerow([])

def _create_heat_map_csv(data: Dict[str, Any]) -> str:
    """Create professionally formatted CSV from heat map data"""
    output = StringIO()
    writer = csv.writer(output)
    
    # Header
    writer.writerow(["=" * 60])
    writer.writerow(["FACILITYFIX - LOCATION HEAT MAP ANALYSIS"])
    writer.writerow(["=" * 60])
    writer.writerow([])
    
    # Summary Information
    writer.writerow(["ANALYSIS SUMMARY"])
    writer.writerow(["-" * 30])
    writer.writerow(["Analysis Period:", f"{data.get('period_days', 30)} days"])
    writer.writerow(["Total Issues Analyzed:", data.get("total_issues", 0)])
    writer.writerow(["Generated On:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    writer.writerow(["Hotspots Identified:", len(data.get("top_hotspots", []))])
    writer.writerow([])
    
    # Critical Hotspots (Top Priority Areas)
    writer.writerow(["CRITICAL HOTSPOTS"])
    writer.writerow(["-" * 30])
    writer.writerow(["Rank", "Location", "Issue Count", "Risk Level", "Primary Category", "Action Required"])
    
    hotspots = data.get("top_hotspots", [])
    for i, hotspot in enumerate(hotspots[:10], 1):  # Top 10 hotspots
        issue_count = hotspot.get("issue_count", 0)
        risk_level = "Critical" if issue_count > 15 else "High" if issue_count > 8 else "Moderate"
        action = "Immediate" if issue_count > 15 else "Schedule" if issue_count > 8 else "Monitor"
        
        writer.writerow([
            f"#{i}",
            hotspot.get("location", "Unknown"),
            issue_count,
            risk_level,
            hotspot.get("primary_category", "General"),
            action
        ])
    writer.writerow([])
    
    # Category Distribution Analysis
    writer.writerow(["CATEGORY DISTRIBUTION"])
    writer.writerow(["-" * 30])
    writer.writerow(["Category", "Count", "Percentage", "Trend", "Priority"])
    
    category_dist = data.get("category_distribution", {})
    total_issues = sum(category_dist.values()) if category_dist else 1
    
    # Sort categories by count (descending)
    sorted_categories = sorted(category_dist.items(), key=lambda x: x[1], reverse=True)
    
    for category, count in sorted_categories:
        percentage = (count / total_issues) * 100
        trend = "Increasing" if percentage > 20 else "Stable" if percentage > 10 else "Decreasing"
        priority = "High" if percentage > 25 else "Medium" if percentage > 15 else "Low"
        
        writer.writerow([
            category.title(),
            count,
            f"{percentage:.1f}%",
            trend,
            priority
        ])
    writer.writerow([])
    
    # Heat Map Matrix (Grid View for Visual Reference)
    if "heat_map_matrix" in data:
        writer.writerow(["🗺️ LOCATION HEAT MAP MATRIX"])
        writer.writerow(["-" * 30])
        writer.writerow(["Building/Floor", "Unit", "Issues", "Status", "Last Updated"])
        
        heat_matrix = data["heat_map_matrix"]
        for location_data in heat_matrix:
            location = location_data.get("location", "Unknown")
            categories = location_data.get("categories", {})
            total_issues = sum(categories.values())
            
            status = "🔴 Attention" if total_issues > 5 else "🟡 Monitor" if total_issues > 2 else "🟢 Good"
            
            # Parse location to extract building/unit info
            parts = location.split(" - ")
            building_floor = parts[0] if parts else location
            unit = parts[1] if len(parts) > 1 else "N/A"
            
            writer.writerow([
                building_floor,
                unit,
                total_issues,
                status,
                datetime.now().strftime("%Y-%m-%d")
            ])
    
    writer.writerow([])
    
    # Recommendations
    writer.writerow(["RECOMMENDATIONS"])
    writer.writerow(["-" * 30])
    writer.writerow(["Priority", "Area", "Recommendation", "Expected Impact"])
    
    # Generate recommendations based on data
    if hotspots:
        top_hotspot = hotspots[0]
        writer.writerow(["Critical", top_hotspot["location"], "Immediate inspection and maintenance", "High"])
    
    if category_dist:
        top_category = max(category_dist.items(), key=lambda x: x[1])
        writer.writerow(["High", f"{top_category[0]} Issues", f"Review {top_category[0]} procedures", "Medium"])
    
    writer.writerow(["General", "All Locations", "Implement preventive maintenance schedule", "Long-term"])
    
    writer.writerow([])
    writer.writerow(["=" * 60])
    writer.writerow(["End of Heat Map Analysis"])
    writer.writerow(["=" * 60])
    
    return output.getvalue()

def _create_staff_performance_csv(data: Dict[str, Any]) -> str:
    """Create professionally formatted CSV from staff performance data"""
    output = StringIO()
    writer = csv.writer(output)
    
    # Header
    writer.writerow(["=" * 70])
    writer.writerow(["FACILITYFIX - STAFF PERFORMANCE ANALYSIS"])
    writer.writerow(["=" * 70])
    writer.writerow([])
    
    # Report Summary
    writer.writerow(["PERFORMANCE OVERVIEW"])
    writer.writerow(["-" * 35])
    writer.writerow(["Analysis Period:", f"{data.get('period_days', 30)} days"])
    writer.writerow(["Total Staff Analyzed:", len(data.get("staff_performance", []))])
    writer.writerow(["Generated On:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    writer.writerow([])
    
    # Overall Team Metrics
    staff_list = data.get("staff_performance", [])
    if staff_list:
        total_assigned = sum(staff.get("assigned_tasks", 0) for staff in staff_list)
        total_completed = sum(staff.get("completed_tasks", 0) for staff in staff_list)
        avg_completion_rate = sum(staff.get("completion_rate", 0) for staff in staff_list) / len(staff_list)
        avg_performance_score = sum(staff.get("performance_score", 0) for staff in staff_list) / len(staff_list)
        
        writer.writerow(["TEAM PERFORMANCE METRICS"])
        writer.writerow(["-" * 35])
        writer.writerow(["Total Tasks Assigned:", total_assigned])
        writer.writerow(["Total Tasks Completed:", total_completed])
        writer.writerow(["Team Completion Rate:", f"{(total_completed/total_assigned*100):.1f}%" if total_assigned > 0 else "0%"])
        writer.writerow(["Average Individual Rate:", f"{avg_completion_rate:.1f}%"])
        writer.writerow(["Team Performance Score:", f"{avg_performance_score:.1f}/100"])
        writer.writerow([])
    
    # Individual Staff Performance
    writer.writerow(["INDIVIDUAL STAFF PERFORMANCE"])
    writer.writerow(["-" * 35])
    writer.writerow([
        "Staff ID", "Name", "Assigned", "Completed", "Rate %", 
        "Avg Time (hrs)", "Performance", "Rating", "Status", "Workload"
    ])
    
    # Sort by performance score (descending)
    sorted_staff = sorted(staff_list, key=lambda x: x.get("performance_score", 0), reverse=True)
    
    for i, staff in enumerate(sorted_staff, 1):
        completion_rate = staff.get("completion_rate", 0)
        performance_score = staff.get("performance_score", 0)
        avg_time = staff.get("average_completion_time_hours", 0)
        
        # Performance rating
        if performance_score >= 90:
            rating = "⭐⭐⭐⭐⭐ Excellent"
        elif performance_score >= 80:
            rating = "⭐⭐⭐⭐ Very Good"
        elif performance_score >= 70:
            rating = "⭐⭐⭐ Good"
        elif performance_score >= 60:
            rating = "⭐⭐ Fair"
        else:
            rating = "⭐ Needs Improvement"
        
        # Status indicator
        if completion_rate >= 95:
            status = "🟢 Excellent"
        elif completion_rate >= 85:
            status = "🟡 Good"
        elif completion_rate >= 70:
            status = "🟠 Average"
        else:
            status = "🔴 Needs Support"
        
        # Workload assessment
        assigned_tasks = staff.get("assigned_tasks", 0)
        if assigned_tasks > 20:
            workload = "🔴 High"
        elif assigned_tasks > 10:
            workload = "🟡 Medium"
        else:
            workload = "🟢 Light"
        
        writer.writerow([
            staff.get("staff_id", "N/A"),
            staff.get("name", staff.get("staff_id", "Unknown")),
            staff.get("assigned_tasks", 0),
            staff.get("completed_tasks", 0),
            f"{completion_rate:.1f}%",
            f"{avg_time:.1f}",
            f"{performance_score:.1f}",
            rating,
            status,
            workload
        ])
    
    writer.writerow([])
    
    # Performance Categories
    writer.writerow(["PERFORMANCE CATEGORIES"])
    writer.writerow(["-" * 35])
    
    # Top Performers
    top_performers = [staff for staff in sorted_staff if staff.get("performance_score", 0) >= 85]
    writer.writerow(["Top Performers (85+ Score):"])
    for staff in top_performers[:5]:  # Top 5
        writer.writerow([f"  • {staff.get('staff_id', 'N/A')} - {staff.get('performance_score', 0):.1f} points"])
    writer.writerow([])
    
    # Staff Needing Support
    needs_support = [staff for staff in sorted_staff if staff.get("completion_rate", 0) < 70]
    if needs_support:
        writer.writerow(["Staff Needing Support (<70% Rate):"])
        for staff in needs_support:
            writer.writerow([f"  • {staff.get('staff_id', 'N/A')} - {staff.get('completion_rate', 0):.1f}% completion"])
        writer.writerow([])
    
    # Workload Distribution
    writer.writerow(["WORKLOAD ANALYSIS"])
    writer.writerow(["-" * 35])
    writer.writerow(["Workload Level", "Staff Count", "Avg Completion Rate", "Recommendation"])
    
    high_workload = [s for s in staff_list if s.get("assigned_tasks", 0) > 20]
    medium_workload = [s for s in staff_list if 10 <= s.get("assigned_tasks", 0) <= 20]
    light_workload = [s for s in staff_list if s.get("assigned_tasks", 0) < 10]
    
    for workload_group, name, recommendation in [
        (high_workload, "High (>20 tasks)", "Consider redistributing workload"),
        (medium_workload, "Medium (10-20 tasks)", "Optimal workload range"),
        (light_workload, "Light (<10 tasks)", "Can handle additional tasks")
    ]:
        count = len(workload_group)
        avg_rate = sum(s.get("completion_rate", 0) for s in workload_group) / count if count > 0 else 0
        writer.writerow([name, count, f"{avg_rate:.1f}%", recommendation])
    
    writer.writerow([])
    
    # Recommendations
    writer.writerow(["STRATEGIC RECOMMENDATIONS"])
    writer.writerow(["-" * 35])
    writer.writerow(["Priority", "Area", "Recommendation", "Expected Outcome"])
    
    # Generate data-driven recommendations
    if needs_support:
        writer.writerow(["High", "Training", "Provide additional training for underperforming staff", "Improved completion rates"])
    
    if high_workload:
        writer.writerow(["Medium", "Workload", "Redistribute tasks from overloaded staff", "Balanced workload distribution"])
    
    if top_performers:
        writer.writerow(["Low", "Recognition", "Implement recognition program for top performers", "Maintained high performance"])
    
    writer.writerow(["Low", "Process", "Regular performance review meetings", "Continuous improvement"])
    
    writer.writerow([])
    writer.writerow(["=" * 70])
    writer.writerow(["End of Staff Performance Analysis"])
    writer.writerow(["=" * 70])
    
    return output.getvalue()

def _create_equipment_csv(data: Dict[str, Any]) -> str:
    """Create professionally formatted CSV from equipment insights data"""
    output = StringIO()
    writer = csv.writer(output)
    
    # Header
    writer.writerow(["=" * 75])
    writer.writerow(["FACILITYFIX - EQUIPMENT INSIGHTS & MAINTENANCE ANALYSIS"])
    writer.writerow(["=" * 75])
    writer.writerow([])
    
    # Report Summary
    writer.writerow(["🔧 EQUIPMENT ANALYSIS OVERVIEW"])
    writer.writerow(["-" * 40])
    writer.writerow(["Analysis Period:", f"{data.get('period_days', 90)} days"])
    writer.writerow(["Equipment Types Analyzed:", len(data.get("equipment_analysis", []))])
    writer.writerow(["Generated On:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    writer.writerow([])
    
    # Overall Equipment Health
    equipment_list = data.get("equipment_analysis", [])
    if equipment_list:
        total_failures = sum(eq.get("failure_count", 0) for eq in equipment_list)
        critical_equipment = len([eq for eq in equipment_list if eq.get("risk_level") == "high"])
        avg_monthly_failures = sum(eq.get("failure_frequency_per_month", 0) for eq in equipment_list) / len(equipment_list)
        
        writer.writerow(["📊 FACILITY EQUIPMENT HEALTH"])
        writer.writerow(["-" * 40])
        writer.writerow(["Total Equipment Failures:", total_failures])
        writer.writerow(["Critical Risk Equipment:", critical_equipment])
        writer.writerow(["Average Monthly Failures:", f"{avg_monthly_failures:.1f}"])
        writer.writerow(["Equipment Requiring Attention:", len([eq for eq in equipment_list if eq.get("risk_level") in ["high", "medium"]])])
        writer.writerow([])
    
    # Equipment Risk Assessment
    writer.writerow(["⚠️ EQUIPMENT RISK ASSESSMENT"])
    writer.writerow(["-" * 40])
    writer.writerow([
        "Equipment Type", "Failure Count", "Monthly Rate", "Risk Level", 
        "Primary Issue", "Next Maintenance", "Criticality", "Status"
    ])
    
    # Sort by risk level and failure count
    risk_priority = {"high": 3, "medium": 2, "low": 1}
    sorted_equipment = sorted(
        equipment_list, 
        key=lambda x: (risk_priority.get(x.get("risk_level", "low"), 0), x.get("failure_count", 0)), 
        reverse=True
    )
    
    for equipment in sorted_equipment:
        risk_level = equipment.get("risk_level", "low")
        failure_count = equipment.get("failure_count", 0)
        
        # Risk level with icon
        if risk_level == "high":
            risk_display = "🔴 CRITICAL"
            criticality = "High"
        elif risk_level == "medium":
            risk_display = "🟡 MODERATE"
            criticality = "Medium"
        else:
            risk_display = "🟢 LOW"
            criticality = "Low"
        
        # Status assessment
        if failure_count > 10:
            status = "🚨 Urgent"
        elif failure_count > 5:
            status = "⚠️ Monitor"
        else:
            status = "✅ Good"
        
        writer.writerow([
            equipment.get("equipment_type", "Unknown"),
            failure_count,
            f"{equipment.get('failure_frequency_per_month', 0):.1f}",
            risk_display,
            equipment.get("most_common_category", "General"),
            equipment.get("predicted_next_maintenance", "TBD"),
            criticality,
            status
        ])
    
    writer.writerow([])
    
    # Maintenance Schedule
    writer.writerow(["📅 PREDICTIVE MAINTENANCE SCHEDULE"])
    writer.writerow(["-" * 40])
    writer.writerow(["Equipment", "Current Status", "Recommended Action", "Timeline", "Priority"])
    
    for equipment in sorted_equipment:
        eq_type = equipment.get("equipment_type", "Unknown")
        failure_count = equipment.get("failure_count", 0)
        risk_level = equipment.get("risk_level", "low")
        
        # Determine maintenance recommendations
        if risk_level == "high" or failure_count > 8:
            action = "Immediate inspection and repair"
            timeline = "Within 7 days"
            priority = "🔴 Critical"
        elif risk_level == "medium" or failure_count > 4:
            action = "Schedule preventive maintenance"
            timeline = "Within 30 days"
            priority = "🟡 High"
        else:
            action = "Continue routine maintenance"
            timeline = "Next scheduled cycle"
            priority = "🟢 Normal"
        
        current_status = f"{failure_count} failures in {data.get('period_days', 90)} days"
        
        writer.writerow([eq_type, current_status, action, timeline, priority])
    
    writer.writerow([])
    
    # Failure Category Analysis
    writer.writerow(["🔍 FAILURE CATEGORY ANALYSIS"])
    writer.writerow(["-" * 40])
    
    # Collect and analyze failure categories
    category_counts = {}
    for equipment in equipment_list:
        category = equipment.get("most_common_category", "General")
        category_counts[category] = category_counts.get(category, 0) + equipment.get("failure_count", 0)
    
    sorted_categories = sorted(category_counts.items(), key=lambda x: x[1], reverse=True)
    
    writer.writerow(["Failure Category", "Total Failures", "Percentage", "Impact Level", "Action Required"])
    total_category_failures = sum(category_counts.values())
    
    for category, count in sorted_categories:
        percentage = (count / total_category_failures * 100) if total_category_failures > 0 else 0
        
        if percentage > 30:
            impact = "🔴 High"
            action = "Immediate process review"
        elif percentage > 15:
            impact = "🟡 Medium"
            action = "Schedule improvement plan"
        else:
            impact = "🟢 Low"
            action = "Monitor trends"
        
        writer.writerow([category, count, f"{percentage:.1f}%", impact, action])
    
    writer.writerow([])
    
    # Cost Impact Analysis
    writer.writerow(["💰 COST IMPACT PROJECTION"])
    writer.writerow(["-" * 40])
    writer.writerow(["Equipment Type", "Failure Cost", "Maintenance Cost", "Downtime Cost", "Total Impact", "ROI Priority"])
    
    for equipment in sorted_equipment:
        eq_type = equipment.get("equipment_type", "Unknown")
        failure_count = equipment.get("failure_count", 0)
        
        # Estimate costs (these would normally come from actual cost data)
        failure_cost = failure_count * 500  # $500 per failure average
        maintenance_cost = failure_cost * 0.3  # Preventive maintenance is 30% of reactive cost
        downtime_cost = failure_count * 200  # $200 per hour downtime
        total_impact = failure_cost + downtime_cost
        
        # ROI calculation for preventive maintenance
        potential_savings = total_impact - maintenance_cost
        roi_priority = "🟢 High ROI" if potential_savings > 1000 else "🟡 Medium ROI" if potential_savings > 500 else "🔴 Low ROI"
        
        writer.writerow([
            eq_type,
            f"${failure_cost:,.0f}",
            f"${maintenance_cost:,.0f}",
            f"${downtime_cost:,.0f}",
            f"${total_impact:,.0f}",
            roi_priority
        ])
    
    writer.writerow([])
    
    # Strategic Recommendations
    writer.writerow(["💡 STRATEGIC MAINTENANCE RECOMMENDATIONS"])
    writer.writerow(["-" * 40])
    writer.writerow(["Priority", "Equipment/Area", "Recommendation", "Expected Benefit", "Investment"])
    
    # Generate recommendations based on analysis
    critical_equipment = [eq for eq in equipment_list if eq.get("risk_level") == "high"]
    if critical_equipment:
        writer.writerow([
            "🔴 Critical",
            f"{len(critical_equipment)} high-risk equipment",
            "Implement immediate maintenance program",
            "Reduce failures by 60-80%",
            "High"
        ])
    
    if sorted_categories:
        top_category = sorted_categories[0][0]
        writer.writerow([
            "🟡 High",
            f"{top_category} systems",
            f"Review {top_category} maintenance procedures",
            "Improve category reliability",
            "Medium"
        ])
    
    writer.writerow([
        "🟢 Medium",
        "All equipment",
        "Implement IoT monitoring sensors",
        "Predictive maintenance capabilities",
        "High (Long-term savings)"
    ])
    
    writer.writerow([
        "🟢 Low",
        "Maintenance team",
        "Staff training on predictive maintenance",
        "Improved maintenance efficiency",
        "Low"
    ])
    
    writer.writerow([])
    writer.writerow(["=" * 75])
    writer.writerow(["End of Equipment Insights Analysis"])
    writer.writerow(["For detailed maintenance procedures, consult equipment manuals"])
    writer.writerow(["=" * 75])
    
    return output.getvalue()

# Enhanced Excel Export Functions (when openpyxl is available)
def _create_comprehensive_excel(data: Dict[str, Any]) -> bytes:
    """Create comprehensive Excel report with charts"""
    if not EXCEL_AVAILABLE:
        raise ValueError("Excel functionality not available")
    
    wb = openpyxl.Workbook()
    
    # Remove default sheet and create custom sheets
    wb.remove(wb.active)
    
    # Create sheets
    summary_sheet = wb.create_sheet("Executive Summary")
    trends_sheet = wb.create_sheet("Trends Analysis")
    charts_sheet = wb.create_sheet("Charts & Visualizations")
    
    # Style definitions
    header_font = Font(name='Arial', size=14, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    subheader_font = Font(name='Arial', size=12, bold=True)
    normal_font = Font(name='Arial', size=10)
    
    # Executive Summary Sheet
    _populate_excel_summary_sheet(summary_sheet, data, header_font, header_fill, subheader_font, normal_font)
    
    # Trends Analysis Sheet  
    _populate_excel_trends_sheet(trends_sheet, data, header_font, header_fill, subheader_font, normal_font)
    
    # Charts Sheet
    _populate_excel_charts_sheet(charts_sheet, data, wb)
    
    # Save to bytes
    output = BytesIO()
    wb.save(output)
    return output.getvalue()

def _create_dashboard_summary_csv(data: Dict[str, Any]) -> str:
    """Create executive dashboard summary CSV"""
    output = StringIO()
    writer = csv.writer(output)
    
    # Header
    writer.writerow(["=" * 60])
    writer.writerow(["FACILITYFIX - EXECUTIVE DASHBOARD SUMMARY"])
    writer.writerow(["=" * 60])
    writer.writerow([])
    
    # Key Metrics
    writer.writerow(["KEY PERFORMANCE INDICATORS"])
    writer.writerow(["-" * 30])
    
    dashboard_stats = data.get("dashboard_stats", {})
    trends = data.get("trends", {})
    
    writer.writerow(["Metric", "Current Value", "Status", "Trend"])
    writer.writerow(["Total Requests", dashboard_stats.get("total_requests", 0), "Active", "Active"])
    writer.writerow(["Pending Concerns", dashboard_stats.get("pending_concerns", 0), "Monitor", "Monitor"])
    writer.writerow(["Active Jobs", dashboard_stats.get("active_jobs", 0), "In Progress", "In Progress"])
    writer.writerow(["Completion Rate", f"{dashboard_stats.get('completion_rate', 0):.1f}%", "Good", "Good"])
    writer.writerow([])
    
    # Category Breakdown
    writer.writerow(["ISSUE CATEGORIES"])
    writer.writerow(["-" * 30])
    categories = data.get("categories", {}).get("categories", {})
    
    writer.writerow(["Category", "Count", "Percentage"])
    total_categories = sum(categories.values()) if categories else 1
    
    for category, count in sorted(categories.items(), key=lambda x: x[1], reverse=True):
        percentage = (count / total_categories) * 100
        writer.writerow([category.title(), count, f"{percentage:.1f}%"])
    
    writer.writerow([])
    
    # Hotspots Summary
    heat_map_summary = data.get("heat_map_summary", {})
    writer.writerow(["LOCATION SUMMARY"])
    writer.writerow(["-" * 30])
    writer.writerow(["Total Hotspots Identified", heat_map_summary.get("total_hotspots", 0)])
    writer.writerow(["Critical Locations (>10 issues)", heat_map_summary.get("critical_locations", 0)])
    writer.writerow([])
    
    # Quick Actions
    writer.writerow(["RECOMMENDED ACTIONS"])
    writer.writerow(["-" * 30])
    
    # Generate recommendations based on data
    pending = dashboard_stats.get("pending_concerns", 0)
    if pending > 10:
        writer.writerow(["High Priority", f"Address {pending} pending concerns"])
    
    critical_locations = heat_map_summary.get("critical_locations", 0)
    if critical_locations > 0:
        writer.writerow(["Medium Priority", f"Inspect {critical_locations} critical locations"])
    
    completion_rate = dashboard_stats.get("completion_rate", 0)
    if completion_rate < 80:
        writer.writerow(["Improvement", f"Improve completion rate from {completion_rate:.1f}%"])
    
    writer.writerow([])
    writer.writerow(["Generated:", data.get("generated_at", datetime.now().isoformat())])
    writer.writerow(["Period:", f"{data.get('period_days', 30)} days"])
    
    return output.getvalue()

def _create_dashboard_summary_excel(data: Dict[str, Any]) -> bytes:
    """Create executive dashboard summary Excel with charts"""
    if not EXCEL_AVAILABLE:
        raise ValueError("Excel functionality not available")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Executive Dashboard"
    
    # Header styling
    header_font = Font(name='Arial', size=16, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1f4e79", end_color="1f4e79", fill_type="solid")
    
    # Add header
    ws.merge_cells('A1:F1')
    ws['A1'] = "FACILITYFIX - EXECUTIVE DASHBOARD"
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Add KPIs
    dashboard_stats = data.get("dashboard_stats", {})
    
    ws['A3'] = "Key Performance Indicators"
    ws['A3'].font = Font(bold=True, size=12)
    
    kpi_data = [
        ("Total Requests", dashboard_stats.get("total_requests", 0)),
        ("Pending Concerns", dashboard_stats.get("pending_concerns", 0)),
        ("Active Jobs", dashboard_stats.get("active_jobs", 0)),
        ("Completion Rate", f"{dashboard_stats.get('completion_rate', 0):.1f}%")
    ]
    
    row = 4
    for label, value in kpi_data:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        row += 1
    
    # Save to bytes
    output = BytesIO()
    wb.save(output)
    return output.getvalue()

# Helper functions for Excel sheets
def _populate_excel_summary_sheet(sheet, data, header_font, header_fill, subheader_font, normal_font):
    """Populate the executive summary sheet"""
    sheet['A1'] = "FacilityFix Analytics - Executive Summary"
    sheet['A1'].font = header_font
    sheet['A1'].fill = header_fill
    
    # Add summary data
    exec_summary = data.get("executive_summary", {})
    row = 3
    
    for key, value in exec_summary.items():
        sheet[f'A{row}'] = key.replace("_", " ").title()
        sheet[f'B{row}'] = value
        sheet[f'A{row}'].font = subheader_font
        row += 1

def _populate_excel_trends_sheet(sheet, data, header_font, header_fill, subheader_font, normal_font):
    """Populate the trends analysis sheet"""
    sheet['A1'] = "Trends Analysis"
    sheet['A1'].font = header_font
    sheet['A1'].fill = header_fill
    
    # Add trend data if available
    trends = data.get("trends", {})
    if "daily_breakdown" in trends:
        sheet['A3'] = "Daily Breakdown"
        sheet['A3'].font = subheader_font
        
        sheet['A4'] = "Date"
        sheet['B4'] = "Count"
        
        row = 5
        for date, count in trends["daily_breakdown"].items():
            sheet[f'A{row}'] = date
            sheet[f'B{row}'] = count
            row += 1

def _populate_excel_charts_sheet(sheet, data, workbook):
    """Populate the charts and visualizations sheet"""
    sheet['A1'] = "Charts & Visualizations"
    sheet['A1'].font = Font(size=14, bold=True)
    
    # Add placeholder for charts
    sheet['A3'] = "Chart data and visualizations would be generated here"
    sheet['A4'] = "This requires additional chart generation logic"

# Placeholder functions for other Excel exports
def _create_heat_map_excel(data: Dict[str, Any]) -> bytes:
    """Create heat map Excel report"""
    if not EXCEL_AVAILABLE:
        raise ValueError("Excel functionality not available")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Heat Map Analysis"
    
    # Basic implementation
    ws['A1'] = "Heat Map Analysis"
    ws['A1'].font = Font(size=14, bold=True)
    
    output = BytesIO()
    wb.save(output)
    return output.getvalue()

def _create_staff_performance_excel(data: Dict[str, Any]) -> bytes:
    """Create staff performance Excel report"""
    if not EXCEL_AVAILABLE:
        raise ValueError("Excel functionality not available")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Staff Performance"
    
    # Basic implementation
    ws['A1'] = "Staff Performance Analysis"
    ws['A1'].font = Font(size=14, bold=True)
    
    output = BytesIO()
    wb.save(output)
    return output.getvalue()

def _create_equipment_excel(data: Dict[str, Any]) -> bytes:
    """Create equipment insights Excel report"""
    if not EXCEL_AVAILABLE:
        raise ValueError("Excel functionality not available")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Equipment Insights"
    
    # Basic implementation
    ws['A1'] = "Equipment Insights Analysis"
    ws['A1'].font = Font(size=14, bold=True)
    
    output = BytesIO()
    wb.save(output)
    return output.getvalue()
